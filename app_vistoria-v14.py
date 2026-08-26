import streamlit as st
import openpyxl
import os
import pandas as pd
from datetime import datetime, time
from PIL import Image

# Suporte ao formato HEIC do iPhone
try:
    from pillow_heif import register_heif_opener
    register_heif_opener()
    HEIF_SUPPORT = True
except ImportError:
    HEIF_SUPPORT = False

# Configuração da página do Streamlit
st.set_page_config(
    page_title="Vistoria de Engenharia Civil",
    page_icon="🚧",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Estilo personalizado com CSS para simular o workflow profissional
st.markdown("""
<style>
    .main-header {
        font-size: 2.2rem;
        color: #2F5496;
        font-weight: bold;
        margin-bottom: 5px;
    }
    .sub-header {
        font-size: 1.1rem;
        color: #555555;
        margin-bottom: 25px;
    }
    .section-title {
        background-color: #4472C4;
        color: white;
        padding: 8px 15px;
        font-weight: bold;
        border-radius: 4px;
        margin-top: 20px;
        margin-bottom: 15px;
    }
    .kpi-card {
        background-color: #F2F2F2;
        border-radius: 8px;
        padding: 15px;
        border-left: 5px solid #2F5496;
        box-shadow: 1px 1px 5px rgba(0,0,0,0.05);
    }
    .kpi-value {
        font-size: 1.8rem;
        font-weight: bold;
        color: #2F5496;
    }
    .kpi-label {
        font-size: 0.9rem;
        color: #666666;
    }
    .wizard-card {
        background-color: #F8F9FA;
        border: 1px solid #4472C4;
        border-radius: 8px;
        padding: 20px;
        margin-bottom: 20px;
    }
    .step-indicator {
        font-size: 1.1rem;
        font-weight: bold;
        color: #2F5496;
        margin-bottom: 15px;
        border-bottom: 2px solid #2F5496;
        padding-bottom: 5px;
    }
</style>
""", unsafe_allow_html=True)

# --- INICIALIZAÇÃO DE ESTADOS DE SESSÃO ---
# Coordenadas gerais
if 'latitude' not in st.session_state:
    st.session_state.latitude = 0.0
if 'longitude' not in st.session_state:
    st.session_state.longitude = 0.0

# Máquina de Estados para o Wizard de Vistoria de Campo
if 'wizard_step' not in st.session_state:
    st.session_state.wizard_step = "setup"  # setup, foto_dano, pergunta_continua

# Retenção de contexto para evitar redigitação em campo
if 'last_ambiente' not in st.session_state:
    st.session_state.last_ambiente = None
if 'last_elemento' not in st.session_state:
    st.session_state.last_elemento = None

# Variáveis temporárias do registro ativo
if 'active_ambiente' not in st.session_state:
    st.session_state.active_ambiente = ""
if 'active_elemento' not in st.session_state:
    st.session_state.active_elemento = ""

# --- DEFINIÇÃO DINÂMICA DOS CAMINHOS ---
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

if os.path.exists("/workspace/artifacts/formulario_vistoria.xlsx") and not os.path.exists(os.path.join(SCRIPT_DIR, "formulario_vistoria.xlsx")):
    FILE_PATH = "/workspace/artifacts/formulario_vistoria.xlsx"
    FOTOS_DIR = "/workspace/scratch/fotos"
else:
    FILE_PATH = os.path.join(SCRIPT_DIR, "formulario_vistoria.xlsx")
    FOTOS_DIR = os.path.join(SCRIPT_DIR, "fotos")

os.makedirs(FOTOS_DIR, exist_ok=True)

def init_excel():
    if not os.path.exists(FILE_PATH):
        st.error(f"❌ O arquivo de modelo não foi encontrado em: **{os.path.abspath(FILE_PATH)}**.\n\n"                 f"Para resolver, certifique-se de que o arquivo **`formulario_vistoria.xlsx`** "                 f"esteja na mesma pasta que este script no seu computador.")
        st.stop()

init_excel()

# Carregar listas de validação do Excel
@st.cache_data
def load_support_lists():
    wb = openpyxl.load_workbook(FILE_PATH, data_only=True)
    ws = wb['Listas_Suporte']
    
    lists = {
        'Pericia': [],
        'Ambiente': [],
        'Elemento': [],
        'Dano': [],
        'FotoClass': []
    }
    
    col_map = {
        1: 'Pericia',
        2: 'Ambiente',
        3: 'Elemento',
        4: 'Dano',
        5: 'FotoClass'
    }
    
    for col_idx, key in col_map.items():
        row = 2
        while True:
            val = ws.cell(row, col_idx).value
            if val is None:
                break
            lists[key].append(str(val))
            row += 1
            
    return lists

support_lists = load_support_lists()

# Função para ler dados de Identificação (Aba 1)
def read_identification_data():
    wb = openpyxl.load_workbook(FILE_PATH, data_only=True)
    ws = wb['1. Identificação e Vistoria']
    
    data = {
        'processo': ws['B5'].value or "",
        'comarca': ws['E5'].value or "",
        'requerente': ws['B6'].value or "",
        'requerido': ws['E6'].value or "",
        'tipo_pericia': ws['B7'].value or support_lists['Pericia'][0],
        'data': ws['B10'].value or datetime.today().date(),
        'hora': ws['E10'].value or "09:00",
        'endereco': ws['B11'].value or "",
        'latitude': ws['B12'].value or 0.0,
        'longitude': ws['E12'].value or 0.0
    }
    
    if isinstance(data['data'], str):
        try:
            data['data'] = datetime.strptime(data['data'], "%Y-%m-%d").date()
        except:
            data['data'] = datetime.today().date()
            
    if isinstance(data['hora'], time):
        data['hora'] = data['hora'].strftime("%H:%M")
        
    return data

# Função para salvar dados de Identificação
def save_identification_data(data):
    wb = openpyxl.load_workbook(FILE_PATH)
    ws = wb['1. Identificação e Vistoria']
    
    ws['B5'] = data['processo']
    ws['E5'] = data['comarca']
    ws['B6'] = data['requerente']
    ws['E6'] = data['requerido']
    ws['B7'] = data['tipo_pericia']
    ws['B10'] = data['data'].strftime("%Y-%m-%d") if isinstance(data['data'], datetime) or hasattr(data['data'], 'strftime') else str(data['data'])
    ws['E10'] = data['hora']
    ws['B11'] = data['endereco']
    ws['B12'] = float(data['latitude']) if data['latitude'] else 0.0
    ws['E12'] = float(data['longitude']) if data['longitude'] else 0.0
    
    wb.save(FILE_PATH)

# Função para ler registros de danos (Aba 2)
def read_damage_records():
    wb = openpyxl.load_workbook(FILE_PATH, data_only=True)
    ws = wb['2. Registro de Danos']
    
    records = []
    row = 8
    while True:
        val_amb = ws.cell(row, 1).value
        val_elem = ws.cell(row, 2).value
        val_dano = ws.cell(row, 3).value
        
        if val_amb is None and val_elem is None and val_dano is None:
            empty_streak = True
            for check_row in range(row, row + 5):
                if any(ws.cell(check_row, c).value is not None for c in range(1, 11)):
                    empty_streak = False
                    row = check_row
                    break
            if empty_streak:
                break
                
        records.append({
            'row_idx': row,
            'ambiente': val_amb or "",
            'elemento': val_elem or "",
            'dano': val_dano or "",
            'descricao': ws.cell(row, 4).value or "",
            'medicoes': ws.cell(row, 5).value or "",
            'observacoes': ws.cell(row, 6).value or "",
            'foto_codigo': ws.cell(row, 7).value or "",
            'foto_class': ws.cell(row, 8).value or "",
            'coordenada': ws.cell(row, 9).value or "",
            'local_exato': ws.cell(row, 10).value or ""
        })
        row += 1
        
    cols = ['row_idx', 'ambiente', 'elemento', 'dano', 'descricao', 'medicoes', 'observacoes', 'foto_codigo', 'foto_class', 'coordenada', 'local_exato']
    if not records:
        return pd.DataFrame(columns=cols)
    return pd.DataFrame(records)

# Função para adicionar registro com duas fotos embutidas (Geral e Detalhe)
def add_damage_record(record):
    wb = openpyxl.load_workbook(FILE_PATH)
    ws = wb['2. Registro de Danos']
    
    row = 8
    while True:
        val_amb = ws.cell(row, 1).value
        val_elem = ws.cell(row, 2).value
        val_dano = ws.cell(row, 3).value
        if val_amb in [None, ""] and val_elem in [None, ""] and val_dano in [None, ""]:
            break
        row += 1
        
    ws.cell(row, 1, record['ambiente'])
    ws.cell(row, 2, record['elemento'])
    ws.cell(row, 3, record['dano'])
    ws.cell(row, 4, record['descricao'])
    ws.cell(row, 5, record['medicoes'])
    ws.cell(row, 6, record['observacoes'])
    
    # Armazena os dois nomes de fotos na coluna G, separados por ponto e vírgula
    foto_comb = f"{record['foto_local_codigo']}; {record['foto_detalhe_codigo']}"
    ws.cell(row, 7, foto_comb)
    ws.cell(row, 8, record['foto_class'])
    ws.cell(row, 9, record['coordenada'])
    ws.cell(row, 10, record['local_exato'])
    
    # Embutir as duas fotos no Excel
    ws.row_dimensions[row].height = 90
    
    # Foto 1: Geral / Local (Ancorada em G)
    if record['foto_local_codigo']:
        foto_l_path = os.path.join(FOTOS_DIR, record['foto_local_codigo'])
        if os.path.exists(foto_l_path):
            try:
                from openpyxl.drawing.image import Image as OpenpyxlImage
                img_embed_l = OpenpyxlImage(foto_l_path)
                img_embed_l.width = 140
                img_embed_l.height = 110
                ws.add_image(img_embed_l, f'G{row}')
            except:
                pass
                
    # Foto 2: Detalhe (Ancorada em H)
    if record['foto_detalhe_codigo']:
        foto_d_path = os.path.join(FOTOS_DIR, record['foto_detalhe_codigo'])
        if os.path.exists(foto_d_path):
            try:
                from openpyxl.drawing.image import Image as OpenpyxlImage
                img_embed_d = OpenpyxlImage(foto_d_path)
                img_embed_d.width = 140
                img_embed_d.height = 110
                ws.add_image(img_embed_d, f'H{row}')
            except:
                pass
                
    wb.save(FILE_PATH)
    return row

def get_kpis():
    df_danos = read_damage_records()
    total_danos_py = len(df_danos[df_danos['ambiente'] != ""])
    total_fotos_py = len(df_danos[(df_danos['foto_codigo'] != "") & (df_danos['foto_codigo'].notna())])
    danos_criticos_py = len(df_danos[df_danos['dano'].isin(["Rachadura", "Corrosão"])])
    
    return {
        'total_danos': total_danos_py,
        'total_fotos': total_fotos_py,
        'danos_criticos': danos_criticos_py
    }

# --- CABEÇALHO ---
st.markdown("<div class='main-header'>🚧 VISTORIA DE CAMPO - ENGENHARIA CIVIL</div>", unsafe_allow_html=True)
st.markdown("<div class='sub-header'>Aplicativo de Coleta Inteligente de Acordo com seu Fluxo Otimizado de Decisão</div>", unsafe_allow_html=True)

kpis = get_kpis()
st.sidebar.markdown("### 📊 Painel Geral de Campo")
col_s1, col_s2 = st.sidebar.columns(2)
with col_s1:
    st.markdown(f"<div class='kpi-card'><div class='kpi-value'>{kpis['total_danos']}</div><div class='kpi-label'>Danos</div></div>", unsafe_allow_html=True)
with col_s2:
    st.markdown(f"<div class='kpi-card' style='border-left-color: #E2C044;'><div class='kpi-value'>{kpis['total_fotos']}</div><div class='kpi-label'>Fotos</div></div>", unsafe_allow_html=True)

st.sidebar.markdown("<br>", unsafe_allow_html=True)
st.markdown(f"""
<div class='kpi-card' style='border-left-color: #C00000; margin-bottom: 20px;'>
    <div class='kpi-value' style='color: #C00000;'>{kpis['danos_criticos']}</div>
    <div class='kpi-label'>Patologias Críticas Identificadas (Corrosão ou Rachaduras)</div>
</div>
""", unsafe_allow_html=True)

# Botão de download da planilha na barra lateral
st.sidebar.markdown("### 💾 Finalização & Exportação")
with open(FILE_PATH, "rb") as f:
    st.sidebar.download_button(
        label="📥 Baixar Planilha com Fotos (.XLSX)",
        data=f,
        file_name="formulario_vistoria_campo.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        help="Baixe a planilha com todas as fotos embutidas e informações sincronizadas."
    )

st.sidebar.info("💡 **Dica de Campo:** O app emite e grava fotos diretamente no Excel de forma proporcional, otimizando a montagem do laudo!")

# Tabs Principais do App
tab_ident, tab_novo_dano, tab_visualizar = st.tabs([
    "📋 1. Identificação e Vistoria", 
    "📸 2. Registrar Novo Dano (Fluxo Otimizado)", 
    "🔍 3. Danos Registrados & Galeria"
])

# --- TAB 1: IDENTIFICAÇÃO E VISTORIA ---
with tab_ident:
    st.markdown("<div class='section-title'>SEÇÃO 1 — IDENTIFICAÇÃO DA PERÍCIA</div>", unsafe_allow_html=True)
    current_ident = read_identification_data()
    
    col1, col2 = st.columns(2)
    with col1:
        processo = st.text_input("Número do Processo", value=current_ident['processo'], placeholder="Ex: 1002345-67.2026.8.26.0100")
        requerente = st.text_input("Requerente", value=current_ident['requerente'], placeholder="Ex: Condomínio Edifício Progresso")
        
        tipo_idx = 0
        if current_ident['tipo_pericia'] in support_lists['Pericia']:
            tipo_idx = support_lists['Pericia'].index(current_ident['tipo_pericia'])
        tipo_pericia = st.selectbox("Tipo de Perícia", options=support_lists['Pericia'], index=tipo_idx)
        
    with col2:
        comarca = st.text_input("Vara / Comarca", value=current_ident['comarca'], placeholder="Ex: 2ª Vara Cível da Capital / SP")
        requerido = st.text_input("Requerido", value=current_ident['requerido'], placeholder="Ex: Construtora Estrutura Forte Ltda.")
    
    st.markdown("<div class='section-title'>SEÇÃO 2 — DADOS DA VISTORIA</div>", unsafe_allow_html=True)
    
    col3, col4 = st.columns(2)
    with col3:
        data_vistoria = st.date_input("Data da Vistoria", value=current_ident['data'])
        endereco = st.text_area("Endereço Completo do Local", value=current_ident['endereco'], placeholder="Ex: Av. Paulista, 1000 - Bela Vista, São Paulo - SP")
        
    with col4:
        hora_str = current_ident['hora']
        try:
            h, m = map(int, hora_str.split(':'))
            hora_time = time(h, m)
        except:
            hora_time = time(9, 0)
        hora_vistoria = st.time_input("Hora da Vistoria", value=hora_time)
        
        if st.session_state.latitude == 0.0 and current_ident['latitude'] != 0.0:
            st.session_state.latitude = float(current_ident['latitude'])
        if st.session_state.longitude == 0.0 and current_ident['longitude'] != 0.0:
            st.session_state.longitude = float(current_ident['longitude'])

        col_lat, col_lon = st.columns(2)
        with col_lat:
            latitude = st.number_input("Latitude (G. Decimais)", value=float(st.session_state.latitude), format="%.6f", min_value=-90.0, max_value=90.0, key="lat_input_widget")
            st.session_state.latitude = latitude
        with col_lon:
            longitude = st.number_input("Longitude (G. Decimais)", value=float(st.session_state.longitude), format="%.6f", min_value=-180.0, max_value=180.0, key="lon_input_widget")
            st.session_state.longitude = longitude

        st.markdown("---")
        st.markdown("### 📡 Captura de Coordenadas de Identificação")
        gps_html = """
        <div style="background-color: #F8F9FA; padding: 15px; border-radius: 8px; border: 1px solid #2F5496; margin-bottom: 10px;">
            <button onclick="getLocation()" style="background-color: #2F5496; color: white; border: none; padding: 10px 18px; border-radius: 4px; cursor: pointer; font-weight: bold; font-size: 0.95rem;">
                📍 Capturar Coordenada da Obra (GPS)
            </button>
            <p id="gps_status" style="margin-top: 10px; font-weight: bold; color: #555; font-size: 0.9rem; margin-bottom: 5px;"></p>
            <p id="gps_coords" style="font-size: 1.1rem; font-weight: bold; color: #2F5496; margin-top: 5px; font-family: monospace;"></p>
        </div>

        <script>
        function getLocation() {
            var status = document.getElementById('gps_status');
            var coords = document.getElementById('gps_coords');
            status.innerHTML = "Buscando sinal de GPS...";
            coords.innerHTML = "";
            
            if (navigator.geolocation) {
                navigator.geolocation.getCurrentPosition(showPosition, showError, {
                    enableHighAccuracy: true,
                    timeout: 10000,
                    maximumAge: 0
                });
            } else {
                status.innerHTML = "❌ Não suportado pelo navegador.";
            }
        }

        function showPosition(position) {
            var status = document.getElementById('gps_status');
            var coords = document.getElementById('gps_coords');
            status.innerHTML = "✅ Localização obtida! Digite nas caixas acima:";
            coords.innerHTML = "Lat: " + position.coords.latitude.toFixed(6) + "<br>Lon: " + position.coords.longitude.toFixed(6);
        }

        function showError(error) {
            var status = document.getElementById('gps_status');
            status.innerHTML = "❌ Bloqueado (requer conexão segura HTTPS ou permissão local).";
        }
        </script>
        """
        st.components.v1.html(gps_html, height=180)

    if st.button("💾 Salvar Identificação & Dados Gerais", type="primary"):
        save_data = {
            'processo': processo,
            'comarca': comarca,
            'requerente': requerente,
            'requerido': requerido,
            'tipo_pericia': tipo_pericia,
            'data': data_vistoria,
            'hora': hora_vistoria.strftime("%H:%M"),
            'endereco': endereco,
            'latitude': latitude,
            'longitude': longitude
        }
        save_identification_data(save_data)
        st.success("✅ Dados da Perícia e Vistoria updated com sucesso no arquivo Excel!")

# --- TAB 2: REGISTRAR NOVO DANO (FLUXO OTIMIZADO - WIZARD) ---
with tab_novo_dano:
    st.markdown("<div class='section-title'>📸 REGISTRO DE PATOLOGIA - WIZARD DE CAMPO</div>", unsafe_allow_html=True)
    
    # Pasos do Wizard
    if st.session_state.wizard_step == "setup":
        st.markdown("""
        <div class='wizard-card'>
            <div class='step-indicator'>PASSO 1: Definir Local de Inspeção (Caracterização Técnica)</div>
        </div>
        """, unsafe_allow_html=True)
        
        # Caso já tenhamos um ambiente registrado na sessão
        if st.session_state.last_ambiente is not None:
            st.info(f"📍 **Ambiente Anterior Ativo:** `{st.session_state.last_ambiente}`")
            mudar_ambiente = st.radio(
                "Você mudou de Ambiente / Setor de vistoria?",
                options=["Não, continuar no mesmo ambiente", "Sim, mudar de ambiente"],
                index=0
            )
            
            if mudar_ambiente == "Sim, mudar de ambiente":
                active_ambiente = st.selectbox("Selecione o Novo Ambiente / Setor", options=support_lists['Ambiente'])
                active_elemento = st.selectbox("Selecione o Elemento Avaliado", options=support_lists['Elemento'])
            else:
                active_ambiente = st.session_state.last_ambiente
                
                st.info(f"🧱 **Elemento Anterior Ativo:** `{st.session_state.last_elemento}`")
                mesmo_elemento = st.radio(
                    "O Elemento Avaliado continua sendo o mesmo?",
                    options=["Sim, manter o elemento anterior", "Não, trocar de elemento"],
                    index=0
                )
                
                if mesmo_elemento == "Não, trocar de elemento":
                    active_elemento = st.selectbox("Selecione o Novo Elemento Avaliado", options=support_lists['Elemento'])
                else:
                    active_elemento = st.session_state.last_elemento
        else:
            st.warning("⚠️ Nenhum registro feito ainda. Defina o primeiro local para iniciar a vistoria.")
            active_ambiente = st.selectbox("Selecione o Ambiente / Setor", options=support_lists['Ambiente'])
            active_elemento = st.selectbox("Selecione o Elemento Avaliado", options=support_lists['Elemento'])
            
        st.session_state.active_ambiente = active_ambiente
        st.session_state.active_elemento = active_elemento
        
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("Avançar para Seção 4 (Fotos & Detalhes) ➡️", type="primary"):
            st.session_state.wizard_step = "foto_dano"
            st.rerun()

    elif st.session_state.wizard_step == "foto_dano":
        st.markdown(f"""
        <div class='wizard-card'>
            <div class='step-indicator'>PASSO 2: Registro de Imagens (Visão Geral & Detalhe)</div>
            <p style='font-size: 1rem; color: #555;'>Inspeção ativa em: <b>{st.session_state.active_ambiente}</b> &gt; <b>{st.session_state.active_elemento}</b></p>
        </div>
        """, unsafe_allow_html=True)
        
        col_f1, col_f2 = st.columns(2)
        with col_f1:
            uploaded_file_local = st.file_uploader("📸 Foto 1: Visão Geral / Local (Obrigatório - Máx 10 MB)", type=["png", "jpg", "jpeg", "heic", "heif"], help="Identifica a região e o tipo de anomalia")
            uploaded_file_detalhe = st.file_uploader("🔍 Foto 2: Detalhe da Anomalia (Obrigatório - Máx 10 MB)", type=["png", "jpg", "jpeg", "heic", "heif"], help="Mostra em close a patologia para descrição técnica e medições")
            foto_class = st.selectbox("Classificação da Foto", options=support_lists['FotoClass'], index=1)
            local_exato = st.text_input("Local Exato (Opcional)", placeholder="Ex: Pilar P4 da ala oeste do 2º subsolo")
            
        with col_f2:
            proximo_id = kpis['total_fotos'] + 1
            foto_codigo_padrao = f"FOTO_{proximo_id:03d}"
            
            # Sincroniza estado para poder resetar
            if 'foto_codigo_text_input' not in st.session_state:
                st.session_state.foto_codigo_text_input = foto_codigo_padrao
                st.session_state.last_proximo_id = proximo_id
            
            if 'last_proximo_id' not in st.session_state or st.session_state.last_proximo_id != proximo_id:
                st.session_state.foto_codigo_text_input = foto_codigo_padrao
                st.session_state.last_proximo_id = proximo_id
                
            col_code_in, col_code_res = st.columns([2, 1])
            with col_code_in:
                # O valor é lido do widget e seu estado é sincronizado
                foto_codigo = st.text_input("Código de Registro da Foto", key="foto_codigo_text_input")
            with col_code_res:
                st.markdown("<div style='height: 28px;'></div>", unsafe_allow_html=True)
                if st.button("🔄 Resetar Nome", help="Reseta o nome da foto de volta para o padrão sequencial automático"):
                    st.session_state.foto_codigo_text_input = f"FOTO_{proximo_id:03d}"
                    st.rerun()
                    
            # Alerta de Duplicidade corrigido para ambas as fotos
            foto_existente = False
            for ext_check in ['.jpg', '.png', '.heic', '.heif', '.jpeg']:
                if os.path.exists(os.path.join(FOTOS_DIR, f"{foto_codigo}_local{ext_check}")) or os.path.exists(os.path.join(FOTOS_DIR, f"{foto_codigo}_detalhe{ext_check}")):
                    foto_existente = True
                    break
            if foto_existente:
                st.warning(f"⚠️ **Nome Duplicado:** Arquivos com o prefixo `{foto_codigo}` já existem na pasta de fotos. Use o botão 'Resetar Nome' ou altere o código para evitar sobreposições.")
                
            dano = st.selectbox("Tipo de Dano / Patologia", options=support_lists['Dano'])
            medicoes = st.text_input("Medições Realizadas", placeholder="Ex: Fissura com abertura de 0.8 mm medida com fissurômetro")

        descricao = st.text_area("Descrição Técnica Detalhada", placeholder="Descreva os sinais visuais, extensão do dano e causas prováveis...")
        observacoes = st.text_area("Observações Complementares", placeholder="Fatores de aceleração de degradação, histórico de reformas...")

        # Mostrar previews horizontais
        if uploaded_file_local is not None or uploaded_file_detalhe is not None:
            st.markdown("**Visualização das Imagens de Campo:**")
            col_p1, col_p2 = st.columns(2)
            with col_p1:
                if uploaded_file_local is not None:
                    ext_l = os.path.splitext(uploaded_file_local.name)[1].lower()
                    if ext_l in ['.heic', '.heif'] and not HEIF_SUPPORT:
                        st.warning("⚠️ Foto 1 do iPhone (HEIC) detectada!")
                    else:
                        try:
                            img_l = Image.open(uploaded_file_local)
                            st.image(img_l, caption=f"{foto_codigo} - 1. Visão Geral / Local", use_container_width=True)
                        except Exception as e:
                            st.error(f"Erro na Foto 1: {e}")
                else:
                    st.info("📸 Faça o upload da Foto 1 (Visão Geral / Local)")
            with col_p2:
                if uploaded_file_detalhe is not None:
                    ext_d = os.path.splitext(uploaded_file_detalhe.name)[1].lower()
                    if ext_d in ['.heic', '.heif'] and not HEIF_SUPPORT:
                        st.warning("⚠️ Foto 2 do iPhone (HEIC) detectada!")
                    else:
                        try:
                            img_d = Image.open(uploaded_file_detalhe)
                            st.image(img_d, caption=f"{foto_codigo} - 2. Detalhe da Anomalia", use_container_width=True)
                        except Exception as e:
                            st.error(f"Erro na Foto 2: {e}")
                else:
                    st.info("🔍 Faça o upload da Foto 2 (Detalhe da Anomalia)")

        st.markdown("<hr>", unsafe_allow_html=True)
        col_b1, col_b2 = st.columns(2)
        with col_b1:
            if st.button("⬅️ Voltar ao Passo 1"):
                st.session_state.wizard_step = "setup"
                st.rerun()
                
        with col_b2:
            if st.button("➕ Gravar Registro de Dano na Planilha", type="primary", use_container_width=True):
                if uploaded_file_local is None or uploaded_file_detalhe is None:
                    st.error("❌ Não é possível gravar! O fluxo de vistoria exige obrigatoriamente ambas as fotos anexadas (Foto 1: Visão Geral / Local E Foto 2: Detalhe da Anomalia) para validar o registro.")
                else:
                    nome_foto_local = ""
                    nome_foto_detalhe = ""
                    
                    # Foto 1: Local
                    ext_l = os.path.splitext(uploaded_file_local.name)[1].lower()
                    filename_l = f"{foto_codigo}_local.jpg"
                    save_path_l = os.path.join(FOTOS_DIR, filename_l)
                    
                    if ext_l in ['.heic', '.heif'] and HEIF_SUPPORT:
                        try:
                            img_heic = Image.open(uploaded_file_local)
                            img_heic.convert("RGB").save(save_path_l, "JPEG")
                            nome_foto_local = filename_l
                        except:
                            filename_l = f"{foto_codigo}_local{ext_l}"
                            save_path_l = os.path.join(FOTOS_DIR, filename_l)
                            with open(save_path_l, "wb") as f:
                                f.write(uploaded_file_local.getbuffer())
                            nome_foto_local = filename_l
                    else:
                        try:
                            img_standard = Image.open(uploaded_file_local)
                            img_standard.convert("RGB").save(save_path_l, "JPEG")
                            nome_foto_local = filename_l
                        except:
                            filename_l = f"{foto_codigo}_local{ext_l}"
                            save_path_l = os.path.join(FOTOS_DIR, filename_l)
                            with open(save_path_l, "wb") as f:
                                f.write(uploaded_file_local.getbuffer())
                            nome_foto_local = filename_l
                            
                    # Foto 2: Detalhe
                    ext_d = os.path.splitext(uploaded_file_detalhe.name)[1].lower()
                    filename_d = f"{foto_codigo}_detalhe.jpg"
                    save_path_d = os.path.join(FOTOS_DIR, filename_d)
                    
                    if ext_d in ['.heic', '.heif'] and HEIF_SUPPORT:
                        try:
                            img_heic = Image.open(uploaded_file_detalhe)
                            img_heic.convert("RGB").save(save_path_d, "JPEG")
                            nome_foto_detalhe = filename_d
                        except:
                            filename_d = f"{foto_codigo}_detalhe{ext_d}"
                            save_path_d = os.path.join(FOTOS_DIR, filename_d)
                            with open(save_path_d, "wb") as f:
                                f.write(uploaded_file_detalhe.getbuffer())
                            nome_foto_detalhe = filename_d
                    else:
                        try:
                            img_standard = Image.open(uploaded_file_detalhe)
                            img_standard.convert("RGB").save(save_path_d, "JPEG")
                            nome_foto_detalhe = filename_d
                        except:
                            filename_d = f"{foto_codigo}_detalhe{ext_d}"
                            save_path_d = os.path.join(FOTOS_DIR, filename_d)
                            with open(save_path_d, "wb") as f:
                                f.write(uploaded_file_detalhe.getbuffer())
                            nome_foto_detalhe = filename_d

                    novo_registro = {
                        'ambiente': st.session_state.active_ambiente,
                        'elemento': st.session_state.active_elemento,
                        'dano': dano,
                        'descricao': descricao,
                        'medicoes': medicoes,
                        'observacoes': observacoes,
                        'foto_local_codigo': nome_foto_local,
                        'foto_detalhe_codigo': nome_foto_detalhe,
                        'foto_class': foto_class,
                        'coordenada': "",
                        'local_exato': local_exato
                    }
                    
                    linha_gravada = add_damage_record(novo_registro)
                    st.success(f"✅ Sucesso! Registro adicionado na Linha {linha_gravada} da planilha Excel (com foto embutida).")
                    
                    st.session_state.last_ambiente = st.session_state.active_ambiente
                    st.session_state.last_elemento = st.session_state.active_elemento
                    
                    st.session_state.wizard_step = "pergunta_continua"
                    st.rerun()

    elif st.session_state.wizard_step == "pergunta_continua":
        st.markdown("""
        <div class='wizard-card' style='text-align: center;'>
            <div class='step-indicator' style='border: none;'>Registro Salvo com Sucesso! 🎉</div>
            <h3 style='color: #2F5496;'>Deseja continuar registrando danos para esta vistoria?</h3>            <p style='color: #666;'>O aplicativo reterá o ambiente e elemento para agilizar suas próximas fotos.</p>        </div>
        """, unsafe_allow_html=True)
        
        col_c1, col_c2 = st.columns(2)
        with col_c1:
            if st.button("🟢 SIM, Registrar Outro Dano", use_container_width=True):
                st.session_state.wizard_step = "setup"
                st.rerun()
        with col_c2:
            if st.button("🔴 NÃO, Finalizar Vistoria", use_container_width=True):
                st.success("Inspeção de campo finalizada! Você pode baixar a planilha atualizada na barra lateral ou visualizar os dados na aba 3.")
                st.session_state.wizard_step = "setup"
                st.rerun()

# --- TAB 3: VISUALIZAR DANOS REGISTRADOS ---
with tab_visualizar:
    st.markdown("<div class='section-title'>LISTA DE ELEMENTOS AVALIADOS ATÉ O MOMENTO</div>", unsafe_allow_html=True)
    
    df_danos = read_damage_records()
    df_danos_clean = df_danos[df_danos['ambiente'] != ""].reset_index(drop=True)
    
    if df_danos_clean.empty:
        st.warning("Nenhum registro de dano inserido nesta vistoria até o momento. Utilize o fluxo para registrar.")
    else:
        df_display = df_danos_clean.copy()
        df_display = df_display.drop(columns=['coordenada'], errors='ignore')
        df_display.columns = [
            'Linha Excel', 'Ambiente / Setor', 'Elemento', 'Dano', 'Descrição Técnica', 
            'Medições', 'Observações', 'Arquivo da Foto', 'Classificação da Foto', 'Local Exato'
        ]
        
        st.dataframe(df_display, use_container_width=True, hide_index=True)
        
        st.markdown("<div class='section-title'>GALERIA FOTOGRÁFICA DE VISTORIA</div>", unsafe_allow_html=True)
        df_com_fotos = df_danos_clean[df_danos_clean['foto_codigo'] != ""].reset_index(drop=True)
        
        if df_com_fotos.empty:
            st.info("Nenhuma imagem de campo cadastrada ainda.")
        else:
            cols_galeria = st.columns(3)
            for idx, row in df_com_fotos.iterrows():
                col_target = cols_galeria[idx % 3]
                foto_nome = row['foto_codigo']
                
                # Trata múltiplos nomes de foto separados por ';'
                nomes_fotos = [n.strip() for n in foto_nome.split(";") if n.strip()]
                
                with col_target:
                    st.markdown(f"**📍 {row['ambiente']} — {row['local_exato'] or 'Área não especificada'}**")
                    st.markdown(f"**Dano:** `{row['dano']}` | **Classe:** `{row['foto_class']}`")
                    
                    if len(nomes_fotos) == 2:
                        col_pic1, col_pic2 = st.columns(2)
                        with col_pic1:
                            f1 = nomes_fotos[0]
                            f1_path = os.path.join(FOTOS_DIR, f1)
                            if os.path.exists(f1_path):
                                try:
                                    img1 = Image.open(f1_path)
                                    st.image(img1, caption="1. Visão Geral", use_container_width=True)
                                except Exception as e:
                                    st.error(f"Erro: {e}")
                            else:
                                st.warning(f"📷 {f1}\n(Não encontrada localmente)")
                        with col_pic2:
                            f2 = nomes_fotos[1]
                            f2_path = os.path.join(FOTOS_DIR, f2)
                            if os.path.exists(f2_path):
                                try:
                                    img2 = Image.open(f2_path)
                                    st.image(img2, caption="2. Detalhe", use_container_width=True)
                                except Exception as e:
                                    st.error(f"Erro: {e}")
                            else:
                                st.warning(f"📷 {f2}\n(Não encontrada localmente)")
                    else:
                        for f in nomes_fotos:
                            f_path = os.path.join(FOTOS_DIR, f)
                            if os.path.exists(f_path):
                                try:
                                    img_single = Image.open(f_path)
                                    st.image(img_single, caption=f, use_container_width=True)
                                except Exception as e:
                                    st.error(f"Erro: {e}")
                            else:
                                st.warning(f"📷 **{f}**: Imagem não encontrada localmente.")
                    st.markdown("---")
